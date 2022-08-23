from openpyxl import load_workbook

book = load_workbook("sf2.xlsx")
sheet = book.active

data = {}

for row in sheet.iter_rows(2):
    if row[2].value is not None: ## valeur vide dans le tableau excel en lui même.
        row[2].value = float(row[2].value) / 100 ##pour le convertir en metre
        data[row[0].value] = row[1].value

book.save("sf2.xlsx")

print(data["Shibasaki"])
