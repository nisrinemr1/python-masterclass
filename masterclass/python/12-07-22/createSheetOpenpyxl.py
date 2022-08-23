from openpyxl import load_workbook

book = load_workbook("first.xlsx")
book.create_sheet("new worksheet")
book.save("first.xlsx")