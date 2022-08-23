from openpyxl import load_workbook

book = load_workbook("Recettes.xlsx")
getIngredient = book["Patates douces fourrés au chedd"]

for row in getIngredient.iter_rows(4):#débuter la lecture à la ligne 4
    if row[0].value is None:
        break


    name = row[0].value
    quantity = row[1].value 
    unity = row[2].value

    if unity is None:
        unity = ""
    print(f"{name}, {quantity} {unity}")


