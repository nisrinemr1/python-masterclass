from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

book = load_workbook("Recettes.xlsx")
sheet = book.active

ingredients = {}

for sheet in book.worksheets:
    for row in sheet.iter_rows(4):
        if row[0].value is None:
            break
 
        name = row[0].value
        quantity = float(row[1].value)
        unity = row[2].value

        if unity is None:
            unity = ""


        if name not in ingredients:
            ingredients[name] = {"quantity": quantity, "unity": unity}

        else:
            ingredients[name]["quantity"] += quantity


book = Workbook()
sheet = book.active

header = ["Name", "Quantité", "Unité"]
sheet.append(header)

ft = Font(underline="single", bold=True)
sheet["A1"].font = ft
sheet["B1"].font = ft
sheet["C1"].font = ft

for ingredient, data in ingredients.items():
    row = [ingredient, data["quantity"]]
    if data["unity"]:
        row.append(data["unity"])
    sheet.append(row)


book.save("listIngredient.xlsx")